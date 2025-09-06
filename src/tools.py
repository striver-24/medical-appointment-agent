import json
import logging
import os
import smtplib
from datetime import datetime, timedelta
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import List, Dict, Any, Optional

import pandas as pd
from apscheduler.schedulers.background import BackgroundScheduler
from dotenv import load_dotenv
from langchain_core.tools import tool

from src.utils import (
    acquire_lock,
    SCHEDULES_FILE_PATH,
    PATIENTS_FILE_PATH,
    INSURANCE_FILE_PATH,
    ADMIN_REPORT_FILE_PATH,
)

# Load environment variables for optional integrations
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- Reminder System Setup ---
scheduler = BackgroundScheduler()
scheduler.start()

# --- Tool Implementations ---

@tool
def lookup_patient(name: str, dob: str) -> Dict[str, Any]:
    """
    Looks up a patient by name and date of birth (YYYY-MM-DD).
    If found, returns their details. Otherwise, indicates a new patient.
    """
    lock_file = f"{PATIENTS_FILE_PATH}.lock"
    try:
        with acquire_lock(lock_file):
            if not os.path.exists(PATIENTS_FILE_PATH):
                return {"status": "new", "message": "Patient database is empty."}

            df = pd.read_csv(PATIENTS_FILE_PATH)
            # Normalize inputs for comparison
            dob_dt = pd.to_datetime(dob).date()
            df['dob'] = pd.to_datetime(df['dob']).dt.date

            patient_record = df[
                (df['name'].str.lower() == name.lower()) &
                (df['dob'] == dob_dt)
            ]

            if not patient_record.empty:
                patient_data = patient_record.iloc[0].to_dict()
                # Ensure dob is a string for JSON serialization
                patient_data['dob'] = str(patient_data['dob'])
                return {
                    "status": "returning",
                    "patient_id": int(patient_data['patient_id']),
                    "data": patient_data
                }
            else:
                return {"status": "new", "message": "No patient found. Please register."}
    except Exception as e:
        logging.error(f"Error in lookup_patient: {e}")
        return {"status": "error", "message": str(e)}

@tool
def register_new_patient(name: str, dob: str, email: str, phone: str) -> Dict[str, Any]:
    """
    Registers a new patient in the system in real-time.
    Generates a new patient_id and saves the details to the patient database.
    """
    lock_file = f"{PATIENTS_FILE_PATH}.lock"
    try:
        with acquire_lock(lock_file):
            if not os.path.exists(PATIENTS_FILE_PATH):
                # Create file with headers if it doesn't exist
                df = pd.DataFrame(columns=['patient_id', 'name', 'dob', 'phone', 'email'])
            else:
                df = pd.read_csv(PATIENTS_FILE_PATH)

            # Generate a new patient ID
            new_patient_id = (df['patient_id'].max() + 1) if not df.empty else 1000

            new_patient_data = {
                "patient_id": new_patient_id,
                "name": name,
                "dob": dob,
                "phone": phone,
                "email": email
            }

            # Append the new patient record
            new_df = pd.DataFrame([new_patient_data])
            new_df.to_csv(PATIENTS_FILE_PATH, mode='a', header=not os.path.exists(PATIENTS_FILE_PATH) or df.empty, index=False)

            return {
                "status": "success",
                "patient_id": new_patient_id,
                "data": new_patient_data
            }
    except Exception as e:
        logging.error(f"Error in register_new_patient: {e}")
        return {"status": "error", "message": str(e)}


@tool
def find_available_slots(doctor: str, duration_min: int) -> Dict[str, Any]:
    """
    Finds the next 3 available consecutive slots for a given doctor and duration.
    """
    lock_file = f"{SCHEDULES_FILE_PATH}.lock"
    try:
        with acquire_lock(lock_file):
            xls = pd.ExcelFile(SCHEDULES_FILE_PATH)
            if doctor not in xls.sheet_names:
                return {"status": "error", "message": f"Doctor '{doctor}' not found. Available doctors are: {', '.join(xls.sheet_names)}."}

            df = pd.read_excel(xls, sheet_name=doctor)
            df['slot_iso'] = pd.to_datetime(df['slot_iso'])
            df = df.sort_values(by='slot_iso')

            available_slots = []
            num_consecutive_slots_needed = duration_min // 15

            # Iterate through the schedule to find consecutive available slots
            for i in range(len(df) - num_consecutive_slots_needed + 1):
                window = df.iloc[i : i + num_consecutive_slots_needed]

                # Check if all slots in the window are 'Available'
                is_window_available = (window['status'] == 'Available').all()
                # Check if the time difference between slots is exactly 15 minutes
                is_consecutive = all(
                    (window.iloc[j+1]['slot_iso'] - window.iloc[j]['slot_iso']) == timedelta(minutes=15)
                    for j in range(len(window) - 1)
                )

                if is_window_available and is_consecutive:
                    start_slot = window.iloc[0]['slot_iso']
                    # Ensure we are checking for slots in the future
                    if start_slot > datetime.now():
                        available_slots.append(start_slot.isoformat())
                        if len(available_slots) == 3:
                            break

            if available_slots:
                return {"status": "success", "slots": available_slots}
            else:
                return {"status": "not_found", "message": "No available slots found for the required duration. Please try another doctor or check back later."}

    except Exception as e:
        logging.error(f"Error in find_available_slots: {e}")
        return {"status": "error", "message": str(e)}

@tool
def book_appointment(slot_iso: str, doctor: str, patient_id: int, patient_name: str) -> Dict[str, Any]:
    """
    Books an appointment by marking the slot as 'Booked' in the schedule.
    This operation is thread-safe using a file lock.
    """
    lock_file = f"{SCHEDULES_FILE_PATH}.lock"
    try:
        with acquire_lock(lock_file):
            xls = pd.ExcelFile(SCHEDULES_FILE_PATH)
            if doctor not in xls.sheet_names:
                return {"status": "error", "message": f"Doctor '{doctor}' not found."}

            df = pd.read_excel(xls, sheet_name=doctor)
            target_slot = pd.to_datetime(slot_iso)

            # Find the row index for the target slot
            slot_index = df.index[pd.to_datetime(df['slot_iso']) == target_slot].tolist()
            if not slot_index:
                return {"status": "error", "message": "Slot not found."}

            idx = slot_index[0]
            if df.loc[idx, 'status'] != 'Available':
                return {"status": "error", "message": "Slot is no longer available."}

            # Update the slot
            df.loc[idx, 'status'] = 'Booked'
            df.loc[idx, 'booked_by'] = f"{patient_name} (ID: {patient_id})"

            # Atomically write changes back to the specific sheet in the Excel file
            all_sheets = {sheet_name: pd.read_excel(xls, sheet_name=sheet_name) for sheet_name in xls.sheet_names}
            all_sheets[doctor] = df

            writer = pd.ExcelWriter(SCHEDULES_FILE_PATH, engine='openpyxl')
            for sheet_name, sheet_df in all_sheets.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.close()


            booking_id = f"BKNG-{patient_id}-{datetime.now().strftime('%Y%m%d%H%M')}"
            return {"status": "success", "booking_id": booking_id, "message": "Appointment booked successfully."}

    except Exception as e:
        logging.error(f"Error in book_appointment: {e}")
        return {"status": "error", "message": str(e)}


@tool
def save_insurance_details(patient_id: int, company: str, member_id: str, group_number: str) -> Dict[str, Any]:
    """Saves patient insurance details to a JSON file."""
    lock_file = f"{INSURANCE_FILE_PATH}.lock"
    try:
        with acquire_lock(lock_file):
            if os.path.exists(INSURANCE_FILE_PATH):
                with open(INSURANCE_FILE_PATH, 'r') as f:
                    data = json.load(f)
            else:
                data = {}

            data[str(patient_id)] = {
                "company": company,
                "member_id": member_id,
                "group_number": group_number,
                "updated_at": datetime.now().isoformat()
            }

            with open(INSURANCE_FILE_PATH, 'w') as f:
                json.dump(data, f, indent=4)

        return {"status": "success", "message": "Insurance details saved."}
    except Exception as e:
        logging.error(f"Error in save_insurance_details: {e}")
        return {"status": "error", "message": str(e)}

@tool
def export_admin_report(booking_id: str, patient_name: str, patient_id: int, dob: str, insurance_status: str, doctor: str, slot_iso: str) -> Dict[str, Any]:
    """Exports a summary of the booking to an admin-facing Excel report."""
    lock_file = f"{ADMIN_REPORT_FILE_PATH}.lock"
    try:
        with acquire_lock(lock_file):
            new_record = pd.DataFrame([{
                "booking_id": booking_id,
                "patient_name": patient_name,
                "patient_id": patient_id,
                "dob": dob,
                "insurance_status": insurance_status,
                "doctor": doctor,
                "appointment_slot": slot_iso,
                "created_at": datetime.now().isoformat()
            }])

            if not os.path.exists(ADMIN_REPORT_FILE_PATH):
                new_record.to_excel(ADMIN_REPORT_FILE_PATH, index=False)
            else:
                # Use openpyxl to append without overwriting
                from openpyxl import load_workbook
                book = load_workbook(ADMIN_REPORT_FILE_PATH)
                writer = pd.ExcelWriter(ADMIN_REPORT_FILE_PATH, engine='openpyxl')
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                startrow = writer.sheets['Sheet1'].max_row
                new_record.to_excel(writer, sheet_name='Sheet1', startrow=startrow, index=False, header=False)
                writer.close()


        return {"status": "success", "message": "Admin report updated."}
    except Exception as e:
        logging.error(f"Error in export_admin_report: {e}")
        return {"status": "error", "message": str(e)}

@tool
def send_email(to_email: str, subject: str, body: str, attachments: Optional[List[str]] = None) -> Dict[str, Any]:
    """
    Sends an email. Uses SMTP environment variables if set, otherwise logs to console.
    """
    use_smtp = all(os.getenv(k) for k in ["SMTP_SERVER", "SMTP_PORT", "SMTP_USERNAME", "SMTP_PASSWORD", "EMAIL_FROM"])

    if not use_smtp:
        logging.info("--- MOCK EMAIL ---")
        logging.info(f"To: {to_email}")
        logging.info(f"Subject: {subject}")
        logging.info(f"Body: {body}")
        if attachments:
            logging.info(f"Attachments: {', '.join(attachments)}")
        logging.info("--------------------")
        return {"status": "success", "message": "Email logged to console (mock)."}

    try:
        msg = MIMEMultipart()
        msg['From'] = os.getenv("EMAIL_FROM")
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        if attachments:
            for file_path in attachments:
                if os.path.exists(file_path):
                    with open(file_path, "rb") as f:
                        part = MIMEApplication(f.read(), Name=os.path.basename(file_path))
                    part['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
                    msg.attach(part)
                else:
                    logging.warning(f"Attachment not found: {file_path}")

        server = smtplib.SMTP(os.getenv("SMTP_SERVER"), int(os.getenv("SMTP_PORT")))
        server.starttls()
        server.login(os.getenv("SMTP_USERNAME"), os.getenv("SMTP_PASSWORD"))
        server.send_message(msg)
        server.quit()
        return {"status": "success", "message": "Email sent successfully."}
    except Exception as e:
        logging.error(f"Error sending email: {e}")
        return {"status": "error", "message": f"Failed to send email: {e}"}

@tool
def send_sms(to_phone: str, message: str) -> Dict[str, Any]:
    """
    Logs an SMS message to the console. Real SMS sending is disabled.
    """
    logging.info("--- SMS Notification (Logged) ---")
    logging.info(f"To: {to_phone}")
    logging.info(f"Message: {message}")
    logging.info("---------------------------------")
    return {"status": "success", "message": "SMS logged to console."}

@tool
def send_intake_form(patient_id: int, booking_id: str, patient_email: str) -> Dict[str, Any]:
    """Sends the new patient intake form via email."""
    form_path = "/mnt/data/New Patient Intake Form (1).pdf"
    if not os.path.exists(form_path):
        logging.error(f"Intake form not found at the required path: {form_path}")
        return {"status": "error", "message": "Intake form PDF not found on server."}

    subject = "Your New Patient Intake Form"
    body = f"""
    Dear Patient,

    Thank you for booking your appointment (ID: {booking_id}).

    Please complete the attached New Patient Intake Form and bring it with you to your appointment.

    We look forward to seeing you.

    Sincerely,
    The Clinic
    """
    return send_email(to_email=patient_email, subject=subject, body=body, attachments=[form_path])

def _reminder_task(message: str, to_phone: str, to_email: str, subject: str):
    """The actual task executed by the scheduler."""
    logging.info(f"Executing reminder job: {subject}")
    send_sms(to_phone, message)
    send_email(to_email, subject, message)

@tool
def schedule_reminder_jobs(booking_id: str, appointment_iso: str, patient_phone: str, patient_email: str) -> Dict[str, Any]:
    """
    Schedules three automated reminders for the appointment.
    1. 48 hours before: Standard reminder.
    2. 24 hours before: Ask about intake form.
    3. 6 hours before: Ask to confirm attendance.
    """
    try:
        appt_time = datetime.fromisoformat(appointment_iso)

        # Reminder 1: 48 hours before
        reminder_1_time = appt_time - timedelta(hours=48)
        if reminder_1_time > datetime.now():
            msg1 = f"Reminder: Your appointment {booking_id} is scheduled for {appt_time.strftime('%A, %B %d at %I:%M %p')}."
            scheduler.add_job(_reminder_task, 'date', run_date=reminder_1_time, args=[msg1, patient_phone, patient_email, "Appointment Reminder"])

        # Reminder 2: 24 hours before
        reminder_2_time = appt_time - timedelta(hours=24)
        if reminder_2_time > datetime.now():
            msg2 = f"Reminder for appointment {booking_id}: Have you completed your new patient intake form? If not, please let us know if you need assistance."
            scheduler.add_job(_reminder_task, 'date', run_date=reminder_2_time, args=[msg2, patient_phone, patient_email, "Action Required: Intake Form"])

        # Reminder 3: 6 hours before
        reminder_3_time = appt_time - timedelta(hours=6)
        if reminder_3_time > datetime.now():
            msg3 = f"Final reminder for appointment {booking_id} today at {appt_time.strftime('%I:%M %p')}. Please reply to this message to confirm your attendance. If you need to cancel, please let us know the reason."
            scheduler.add_job(_reminder_task, 'date', run_date=reminder_3_time, args=[msg3, patient_phone, patient_email, "Action Required: Confirm Your Appointment"])

        return {"status": "success", "message": "Reminders scheduled successfully."}
    except Exception as e:
        logging.error(f"Error scheduling reminders: {e}")
        return {"status": "error", "message": str(e)}

# List of all tools for the agent
all_tools = [
    lookup_patient,
    register_new_patient,
    find_available_slots,
    book_appointment,
    save_insurance_details,
    export_admin_report,
    send_email,
    send_sms,
    send_intake_form,
    schedule_reminder_jobs,
]